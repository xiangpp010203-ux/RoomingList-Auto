import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
import copy
from io import BytesIO
import os

st.set_page_config(page_title="OP 救星：分房表轉換神器", page_icon="🏨")
st.title("🏨 OP 救星：SCM_Rooming List 自動轉換神器")
st.write("終極完美版：支援無限人數自動擴充、結尾動態公式與格式重組！")

# ==========================================
# ★ 輔助函式區
# ==========================================
def copy_style(source_cell, target_cell):
    """格式複製：完美複製框線、字體、背景色"""
    if source_cell.has_style:
        target_cell.font = copy.copy(source_cell.font)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.number_format = copy.copy(source_cell.number_format)
        target_cell.alignment = copy.copy(source_cell.alignment)

def clean_str(val):
    """資料清洗：徹底消滅 nan 與隱形空白鍵"""
    if pd.isna(val): return ""
    s = str(val).strip()
    if s.lower() == 'nan' or s == '': return ""
    return s

# ==========================================
# 主程式區塊
# ==========================================
uploaded_file_B = st.file_uploader("請上傳【內部系統分房表 (檔案B)】支援 Excel 或 CSV", type=["xlsx", "csv"])

if uploaded_file_B is not None:
    st.success(f"成功讀取檔案：{uploaded_file_B.name}！正在啟動排版引擎...")
    
    try:
        base_name, ext = os.path.splitext(uploaded_file_B.name)
        output_filename = f"{base_name}_RoomingList.xlsx"
        
        # 智慧尋找標題列
        if uploaded_file_B.name.endswith('.csv'):
            df_temp = pd.read_csv(uploaded_file_B, header=None)
            header_idx = 0
            for i in range(min(5, len(df_temp))):
                if '房號' in str(df_temp.iloc[i].values) or '英文姓名' in str(df_temp.iloc[i].values):
                    header_idx = i
                    break
            uploaded_file_B.seek(0) 
            df_B = pd.read_csv(uploaded_file_B, header=header_idx)
        else: 
            df_temp = pd.read_excel(uploaded_file_B, header=None, engine='openpyxl')
            header_idx = 0
            for i in range(min(5, len(df_temp))):
                if '房號' in str(df_temp.iloc[i].values) or '英文姓名' in str(df_temp.iloc[i].values):
                    header_idx = i
                    break
            uploaded_file_B.seek(0)
            df_B = pd.read_excel(uploaded_file_B, header=header_idx, engine='openpyxl')

        df_B.columns = df_B.columns.str.strip()
        
        wb = openpyxl.load_workbook("Template.xlsx")
        sheet = wb.active 
        
        start_row = 14 
        
        # 定位底部的「總房晚」列
        summary_row = None
        for r in range(start_row, sheet.max_row + 1):
            for c in range(1, 4):
                val = str(sheet.cell(row=r, column=c).value or "")
                if "總房晚" in val or "Total Room Nights" in val:
                    summary_row = r
                    break
            if summary_row:
                break
                
        # 解開旅客區塊原有的合併儲存格
        end_unmerge_row = summary_row - 1 if summary_row else sheet.max_row
        ranges_to_unmerge = [str(r) for r in sheet.merged_cells.ranges if r.min_row >= start_row and r.max_row <= end_unmerge_row]
        for r_str in ranges_to_unmerge:
            sheet.unmerge_cells(r_str)
            
        current_row = start_row
        df_B_valid = df_B.dropna(subset=['房號'])
        
        room_numbers = []
        for r in df_B_valid['房號']:
            if pd.notna(r) and r not in room_numbers:
                room_numbers.append(r)
                
        for room_no in room_numbers:
            room_group = df_B_valid[df_B_valid['房號'] == room_no]
            passengers = room_group.to_dict('records')
            num_people = len(passengers)
            
            rows_to_occupy = max(2, num_people)
            start_room_row = current_row 
            
            room_remarks = []
            for p in passengers:
                r = clean_str(p.get('備註'))
                if r: room_remarks.append(r)
            room_remark_str = " / ".join(room_remarks)
            
            for i in range(rows_to_occupy):
                # 智慧擴充機制
                need_insert = False
                if i >= 2:
                    need_insert = True
                elif summary_row and current_row >= summary_row:
                    need_insert = True
                    
                if need_insert:
                    sheet.insert_rows(current_row)
                    if summary_row and current_row <= summary_row:
                        summary_row += 1
                    
                    if sheet.row_dimensions[current_row - 1].height:
                        sheet.row_dimensions[current_row].height = sheet.row_dimensions[current_row - 1].height
                    for col in range(1, sheet.max_column + 1):
                        copy_style(sheet.cell(row=current_row - 1, column=col), 
                                   sheet.cell(row=current_row, column=col))
                                   
                # 填寫真實旅客資料
                if i < num_people:
                    row = passengers[i]
                    
                    eng_name_raw = clean_str(row.get('英文姓名'))
                    title, last_name, first_name = "", "", ""
                    if "MR " in eng_name_raw: title, eng_name_raw = "Mr", eng_name_raw.replace("MR ", "")
                    elif "MS " in eng_name_raw: title, eng_name_raw = "Ms", eng_name_raw.replace("MS ", "")
                    elif "MISS " in eng_name_raw: title, eng_name_raw = "Miss", eng_name_raw.replace("MISS ", "")
                    elif "MSTR " in eng_name_raw: title, eng_name_raw = "Mstr", eng_name_raw.replace("MSTR ", "")
                    
                    if "/" in eng_name_raw:
                        last_name, first_name = eng_name_raw.split("/", 1)
                        last_name, first_name = last_name.strip(), first_name.strip()
                    else:
                        last_name = eng_name_raw

                    dob_raw = clean_str(row.get('生日'))
                    if dob_raw.endswith('.0'): dob_raw = dob_raw[:-2]
                    dob_formatted = f"{dob_raw[0:4]}/{dob_raw[4:6]}/{dob_raw[6:8]}" if (len(dob_raw) == 8 and dob_raw.isdigit()) else dob_raw
                        
                    cht_name = clean_str(row.get('中文姓名'))
                    cht_last = cht_name[0] if len(cht_name) >= 1 else ""
                    cht_first = cht_name[1:] if len(cht_name) >= 2 else ""
                        
                    room_raw = clean_str(room_no)
                    if room_raw.endswith('.0'): room_raw = room_raw[:-2]
                    
                    no_raw = clean_str(row.get('No'))
                    if no_raw.endswith('.0'): no_raw = no_raw[:-2]
                        
                    passport = clean_str(row.get('護照號碼'))
                    if passport.endswith('.0'): passport = passport[:-2]

                    sheet.cell(row=current_row, column=1).value = room_raw if i == 0 else ""
                    sheet.cell(row=current_row, column=2).value = no_raw
                    sheet.cell(row=current_row, column=3).value = title
                    sheet.cell(row=current_row, column=4).value = last_name
                    sheet.cell(row=current_row, column=5).value = first_name
                    sheet.cell(row=current_row, column=6).value = cht_last
                    sheet.cell(row=current_row, column=7).value = cht_first
                    sheet.cell(row=current_row, column=10).value = passport
                    sheet.cell(row=current_row, column=11).value = dob_formatted
                    
                    if i == 0:
                        sheet.cell(row=current_row, column=21).value = room_remark_str
                
                # 單人房空白列清空
                else:
                    for c in range(2, 12):
                        sheet.cell(row=current_row, column=c).value = ""
                
                current_row += 1
                
            end_room_row = current_row - 1 
            
            # 合併 A 欄與 L~V 欄
            if start_room_row < end_room_row:
                cols_to_merge = [1] + list(range(12, 23))
                for col in cols_to_merge:
                    sheet.merge_cells(start_row=start_room_row, start_column=col, end_row=end_room_row, end_column=col)
                    top_cell = sheet.cell(row=start_room_row, column=col)
                    top_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # ==========================================
        # 刪除多餘列，對齊總計區塊
        # ==========================================
        if summary_row and current_row < summary_row:
            rows_to_delete = summary_row - current_row
            sheet.delete_rows(current_row, rows_to_delete)
        
        # ==========================================
        # ★ 最新優化：精準重構「總房晚」列的格式與動態公式
        # ==========================================
        final_summary_row = current_row
        
        # 1. 為了安全，先解除最後一列原本的所有合併狀態
        ranges_to_unmerge = [str(r) for r in sheet.merged_cells.ranges if r.min_row <= final_summary_row <= r.max_row]
        for r_str in ranges_to_unmerge:
            sheet.unmerge_cells(r_str)
            
        # 2. A~Q欄 (1~17)：合併並靠右對齊
        sheet.merge_cells(start_row=final_summary_row, start_column=1, end_row=final_summary_row, end_column=17)
        cell_total = sheet.cell(row=final_summary_row, column=1)
        cell_total.value = "總房晚 Total Room Nights "
        # 保持原本字體(通常是粗體)，並設定靠右置中
        cell_total.alignment = Alignment(horizontal='right', vertical='center')
        
        # 3. R欄 (18)：注入動態 SUM 公式
        cell_r = sheet.cell(row=final_summary_row, column=18)
        # 公式範圍從 14 列一直加總到 最後一列 的上一列
        cell_r.value = f"=SUM(R14:R{final_summary_row - 1})"
        cell_r.alignment = Alignment(horizontal='center', vertical='center')
        
        # 4. S~V欄 (19~22)：合併，維持原本的底色 (模板自帶)
        sheet.merge_cells(start_row=final_summary_row, start_column=19, end_row=final_summary_row, end_column=22)
                
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        st.balloons() 
        st.success(f"🎉 結尾優化完成！總房晚公式與 A~Q、S~V 欄位完美重組！")
        
        st.download_button(
            label=f"📥 下載終極名單 ({output_filename})",
            data=output,
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        st.error(f"系統發生異常：\n{e}")